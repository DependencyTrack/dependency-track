import openpyxl
import json
import warnings
from configparser import ConfigParser
import requests

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
headers = {'X-Api-Key': 'key', 'accept': 'application/json', 'content-type': 'application/json'}
analysisRequest = { 'project': '', 'component': '', 'vulnerability': '', 'analysisState': '', 'analysisJustification': '', 'analysisResponse': '', 'analysisDetails': '', 'comment': '', 'suppressed': 'false' }

# instantiate
config = ConfigParser(delimiters=('='))
config.optionxform = lambda option: option  # preserve case for letters

# parse excel.ini file
config.read('excel.ini')
excelFile = config.get('ExcelData', 'file')
md = config.get('ExcelData', 'metadata')
vu = config.get('ExcelData', 'vulnerabilities')
headers['X-Api-Key'] = config.get('DTData', 'api_key')
base_url = config.get('DTData', 'base_url')

# Load the workbook
wb = openpyxl.load_workbook(excelFile)

# Select the active worksheet
metadata = wb[md]

# build VEX Metadata from Excel
prjName = metadata["B1"].value
prjVersion = str(metadata["B2"].value)

# search for project & version name & version are the unique project ID
# curl -X GET "http://localhost:8081/api/v1/project/lookup?name=Amt24&version=1" -H  "accept: application/json" -H  "X-Api-Key: <key>"
# response: "uuid"
# DT API call
dt_url = base_url + 'api/v1/project/lookup?name=' + prjName + '&version=' + prjVersion
r = requests.get(dt_url, headers=headers)
r.raise_for_status()
json_response = r.json()
prjUUID = json_response["uuid"]

# get list of vulnerabilities from Excel
vulns = wb[vu]

for i in range(2, vulns.max_row+1):
    if not (vulns.cell(row=i, column = 1).value is None ):

        # search for affected component & version
        # curl -X GET "http://localhost:8081/api/v1/component/identity?name=LatencyUtils&version=2.0.3&project=03cc011f-25c7-45c9-841c-b079c126acf8" -H  "accept: application/json" -H  "X-Api-Key: <key>"
        # response: list of "uuid"
        dt_url = base_url + 'api/v1/component/identity?name=' + vulns.cell(row=i, column = 1).value + '&version=' + str(vulns.cell(row=i, column = 2).value) + '&project=' + prjUUID
        r = requests.get(dt_url, headers=headers)
        r.raise_for_status()
        components = r.json()

        for j in range(len(components)):
            # gets all vulnerabilities of this component
            # http://localhost:8081/api/v1/vulnerability/component/2b17edae-605a-4b09-8009-df960f9f6fed
            dt_url = base_url + 'api/v1/vulnerability/component/' + components[j]['uuid']
            r = requests.get(dt_url, headers=headers)
            r.raise_for_status()
            vulnerabilities = r.json()
            vulID = vulns.cell(row=i, column = 3).value
            for k in range(len(vulnerabilities)):
                if vulnerabilities[k]['vulnId'] == vulID:
                    # print('VulnerabilityID (UUID): {}'.format(vulnerabilities[k]['uuid']))
                    analysisRequest['project'] = prjUUID
                    analysisRequest['component'] = components[j]['uuid']
                    analysisRequest['vulnerability'] = vulnerabilities[k]['uuid']
                    analysisRequest['analysisState'] = str(vulns.cell(row=i, column = 4).value)
                    analysisRequest['analysisJustification'] = str(vulns.cell(row=i, column = 5).value)
                    analysisRequest['analysisResponse'] = str(vulns.cell(row=i, column = 6).value)
                    analysisRequest['analysisDetails'] = str(vulns.cell(row=i, column = 7).value)
                    analysisRequest['comment'] = ''
                    analysisRequest['suppressed'] = 'false'
                
                    # API call
                    dt_url = base_url + 'api/v1/analysis'
                    r = requests.put(dt_url, data = json.dumps(analysisRequest), headers=headers)
                    r.raise_for_status()
                    rx = r.json()  
                    print(rx)

wb.close
