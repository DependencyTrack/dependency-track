import openpyxl
import json
import warnings
from datetime import datetime
from configparser import ConfigParser

import pytz
import uuid
import copy
import requests
import base64

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
headers = {'X-Api-Key': 'key', 'accept': 'application/json', 'content-type': 'application/json'}

# instantiate
config = ConfigParser(delimiters=('='))
config.optionxform = lambda option: option  # preserve case for letters

# parse excel.ini file
config.read('excel.ini')
excelFile = config.get('ExcelData', 'file')
md = config.get('ExcelData', 'metadata')
vu = config.get('ExcelData', 'vulnerabilities')
tz = config.get('ExcelData', 'timezone')
jsonFile = config.get('Vex', 'file')
headers['X-Api-Key'] = config.get('DTData', 'api_key')
base_url = config.get('DTData', 'base_url')

# load dummy VEX as a cheat sheet
with open(jsonFile) as json_file:
    json_data = json.load(json_file)
json_file.close

# get json Metadata/Component structure from cheat sheet
comp = json_data['metadata']['component']

# Load the workbook
wb = openpyxl.load_workbook(excelFile)

# Select the active worksheet
metadata = wb[md]

# build VEX Metadata from Excel
comp['name'] = metadata["B1"].value
comp['version'] = metadata["B2"].value
comp['type'] = metadata["B3"].value

# create TimeStamp from date and time values
dtx = metadata["B4"].value
ttx = metadata["B5"].value

# convert local time to utc
local = pytz.timezone(tz)
ts = (datetime.combine(dtx, ttx))
local_dt = local.localize(ts)
utc_dt = local_dt.astimezone(pytz.utc)
json_data['metadata']['timestamp'] = utc_dt.strftime("%Y-%m-%dT%H:%M:%SZ")

# generate bom-ref
bom_ref = str(uuid.uuid4())
comp['bom-ref'] = bom_ref

# get list of vulnerabilities from Excel
vulns = wb[vu]

# build VEX vulnerabilities from Excel
# get json Vulnerabilities structure from cheat sheet
vulx = copy.deepcopy(json_data['vulnerabilities'][0])

# copy the values of Excel cells to jason structure
print(vulns.max_row+1)
for i in range(2, vulns.max_row+1):
    if not (vulns.cell(row=i, column = 1).value is None ):
        vulx['id'] = vulns.cell(row=i, column = 1).value
        vulx['source']['name'] = 'NVD'
        vulx['source']['url'] = 'https://nvd.nist.gov/vuln/detail/' + vulns.cell(row=i, column = 1).value
        vulx['analysis']['state'] = str(vulns.cell(row=i, column = 3).value).lower()
        vulx['analysis']['justification'] = str(vulns.cell(row=i, column = 4).value).lower()
        vulx['analysis']['response'] = str(vulns.cell(row=i, column = 5).value).lower()
        vulx['analysis']['detail'] = str(vulns.cell(row=i, column = 6).value).lower()
        vulx['affects'] = bom_ref
        print(vulx)
        if (i > 2):
            json_data['vulnerabilities'].append(vulx)
        json_data['vulnerabilities'][i - 2] = copy.deepcopy(vulx)

wb.close

# for debug, write the VEX to disk
with open('data.json', 'w', encoding='utf-8') as f:
    json.dump(json_data, f, ensure_ascii=False, indent=2)
f.close

# encode VEX data to BASE64
vex_str = json.dumps(json_data)
vex = (base64.b64encode(vex_str.encode('utf-8'))).decode('ascii')

# creeate request body for DT API call
bodyVEX = {'projectName': '','projectVersion': '', 'vex': ''}
bodyVEX['projectName'] = comp['name']
bodyVEX['projectVersion'] = comp['version']
bodyVEX['vex'] = vex

# DT API call
dt_url = base_url + 'api/v1/vex'
r = requests.put(dt_url, data=json.dumps(bodyVEX), headers=headers)
r.raise_for_status()
json_response = r.json()

print(json_response)
